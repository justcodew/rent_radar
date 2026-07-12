# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# 多账号 —— 账号管理:Excel 批量导入/导出 + CRUD 便捷方法
#
# Excel 模板列(platform, account_id, nickname, cookies, user_agent, proxy_http, proxy_https)
# 参见 account/templates/accounts_template.xlsx

import csv
import json
import os
from pathlib import Path
from typing import List

from app.services.crawler.account import store
from app.services.crawler.account.types import AccountInfo

TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"


async def import_from_csv(csv_path: str) -> int:
    """从 CSV 导入账号,返回导入条数

    CSV 表头(顺序无关,大小写不敏感):
        platform, account_id, nickname, cookies, user_agent, proxy_http, proxy_https
    """
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"账号 CSV 不存在: {csv_path}")

    count = 0
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = {k.lower().strip(): k for k in (reader.fieldnames or [])}
        for row in reader:
            get = lambda key: row.get(headers.get(key, ""), "").strip()
            proxy_cfg = {}
            ph = get("proxy_http")
            if ph:
                proxy_cfg["http"] = ph
            phs = get("proxy_https")
            if phs:
                proxy_cfg["https"] = phs
            info = AccountInfo(
                account_id=get("account_id") or f"acc_{count}",
                platform=get("platform") or "xhs",
                cookies=get("cookies"),
                user_agent=get("user_agent"),
                nickname=get("nickname"),
                proxy_config=proxy_cfg or None,
            )
            await store.add_account(info)
            count += 1
    return count


async def import_from_excel(xlsx_path: str) -> int:
    """从 Excel 导入账号(依赖 openpyxl),返回导入条数"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("Excel 导入需要 openpyxl: uv pip install openpyxl") from e

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    header = [str(c).lower().strip() if c else "" for c in rows[0]]

    def col(row, key):
        if key in header:
            v = row[header.index(key)]
            return str(v).strip() if v is not None else ""
        return ""

    count = 0
    for row in rows[1:]:
        if not any(row):
            continue
        proxy_cfg = {}
        if col(row, "proxy_http"):
            proxy_cfg["http"] = col(row, "proxy_http")
        if col(row, "proxy_https"):
            proxy_cfg["https"] = col(row, "proxy_https")
        info = AccountInfo(
            account_id=col(row, "account_id") or f"acc_{count}",
            platform=col(row, "platform") or "xhs",
            cookies=col(row, "cookies"),
            user_agent=col(row, "user_agent"),
            nickname=col(row, "nickname"),
            proxy_config=proxy_cfg or None,
        )
        await store.add_account(info)
        count += 1
    return count


async def export_to_csv(csv_path: str, platform: str = "") -> int:
    """导出账号到 CSV"""
    infos = await store.list_accounts(platform=platform or None)
    fields = ["platform", "account_id", "nickname", "cookies", "user_agent", "proxy_http", "proxy_https"]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for info in infos:
            w.writerow({
                "platform": info.platform,
                "account_id": info.account_id,
                "nickname": info.nickname,
                "cookies": info.cookies,
                "user_agent": info.user_agent,
                "proxy_http": info.proxy_config.get("http", ""),
                "proxy_https": info.proxy_config.get("https", ""),
            })
    return len(infos)


def write_template_csv(out_path: str) -> None:
    """生成导入模板 CSV"""
    fields = ["platform", "account_id", "nickname", "cookies", "user_agent", "proxy_http", "proxy_https"]
    sample = ["xhs", "acc001", "账号备注A", "web_session=xxx; a1=yyy",
              "Mozilla/5.0 ...", "", ""]
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(fields)
        w.writerow(sample)
